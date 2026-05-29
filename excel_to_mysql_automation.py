"""
Excel to MySQL Automation Script
Monitors Outlook emails for Excel attachments and syncs data to MySQL database
Production-ready with error handling, logging, and retry logic
"""

import os
import sys
import logging
import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Tuple
import traceback

# Third-party imports
from dotenv import load_dotenv
import mysql.connector
from mysql.connector import Error as MySQLError, pooling
from exchangelib import Account, Credentials, FileAttachment, Configuration
from exchangelib.errors import AutoDiscoverFailed, UnauthorizedError
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Load environment variables
load_dotenv()

# ============================================================================
# CONFIGURATION
# ============================================================================

class Config:
    """Configuration manager"""
    
    # Outlook/Microsoft 365
    OUTLOOK_EMAIL = os.getenv('OUTLOOK_EMAIL')
    OUTLOOK_PASSWORD = os.getenv('OUTLOOK_PASSWORD')  # Use app password for Office 365
    
    # MySQL
    MYSQL_HOST = os.getenv('MYSQL_HOST', 'localhost')
    MYSQL_USER = os.getenv('MYSQL_USER')
    MYSQL_PASSWORD = os.getenv('MYSQL_PASSWORD')
    MYSQL_DATABASE = os.getenv('MYSQL_DATABASE')
    MYSQL_PORT = int(os.getenv('MYSQL_PORT', 3306))
    
    # Email notifications
    NOTIFY_EMAIL = os.getenv('NOTIFY_EMAIL')  # Email to send reports to
    NOTIFY_ON_SUCCESS = os.getenv('NOTIFY_ON_SUCCESS', 'False').lower() == 'true'
    NOTIFY_ON_ERROR = os.getenv('NOTIFY_ON_ERROR', 'True').lower() == 'true'
    
    # Automation settings
    DAYS_BACK = int(os.getenv('DAYS_BACK', 1))
    TABLE_NAME = os.getenv('TABLE_NAME', 'excel_data')
    TEMP_DIR = os.getenv('TEMP_DIR', './temp_files')
    
    # Retry settings
    MAX_RETRIES = int(os.getenv('MAX_RETRIES', 3))
    RETRY_DELAY = int(os.getenv('RETRY_DELAY', 5))  # seconds
    
    @staticmethod
    def validate():
        """Validate required configuration"""
        required = ['OUTLOOK_EMAIL', 'OUTLOOK_PASSWORD', 'MYSQL_USER', 'MYSQL_PASSWORD', 'MYSQL_DATABASE']
        missing = [key for key in required if not getattr(Config, key)]
        if missing:
            raise ValueError(f"Missing required configuration: {', '.join(missing)}")


# ============================================================================
# LOGGING SETUP
# ============================================================================

def setup_logging(log_file='excel_to_mysql.log'):
    """Configure logging with both file and console output"""
    
    log_dir = Path('logs')
    log_dir.mkdir(exist_ok=True)
    
    logger = logging.getLogger('ExcelToMySQL')
    logger.setLevel(logging.DEBUG)
    
    # File handler
    file_handler = logging.FileHandler(log_dir / log_file)
    file_handler.setLevel(logging.DEBUG)
    
    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    
    # Formatter
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger


logger = setup_logging()


# ============================================================================
# EMAIL NOTIFICATIONS
# ============================================================================

class EmailNotifier:
    """Send email notifications"""
    
    @staticmethod
    def send_notification(subject: str, body: str, is_error: bool = False):
        """Send email notification"""
        
        if not Config.NOTIFY_EMAIL:
            return
        
        if is_error and not Config.NOTIFY_ON_ERROR:
            return
        if not is_error and not Config.NOTIFY_ON_SUCCESS:
            return
        
        try:
            # This uses Outlook's SMTP server
            # For Office 365, use: smtp.office365.com:587
            msg = MIMEMultipart()
            msg['From'] = Config.OUTLOOK_EMAIL
            msg['To'] = Config.NOTIFY_EMAIL
            msg['Subject'] = subject
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Note: This requires app password for Office 365
            # In production, use OAuth2 instead
            with smtplib.SMTP('smtp.office365.com', 587) as server:
                server.starttls()
                server.login(Config.OUTLOOK_EMAIL, Config.OUTLOOK_PASSWORD)
                server.send_message(msg)
            
            logger.info(f"Notification email sent to {Config.NOTIFY_EMAIL}")
        except Exception as e:
            logger.warning(f"Failed to send notification email: {e}")


# ============================================================================
# OUTLOOK/EXCHANGE INTEGRATION
# ============================================================================

class OutlookIntegration:
    """Handle Outlook/Exchange connectivity and operations"""
    
    def __init__(self, email: str, password: str, max_retries: int = 3):
        self.email = email
        self.password = password
        self.max_retries = max_retries
        self.account = None
    
    def connect(self) -> bool:
        """Connect to Outlook with retry logic"""
        
        for attempt in range(self.max_retries):
            try:
                logger.info(f"Connecting to Outlook... (Attempt {attempt + 1}/{self.max_retries})")
                
                credentials = Credentials(self.email, self.password)
                
                # For Office 365, explicitly set the server
                config = Configuration(
                    server='outlook.office365.com',
                    credentials=credentials
                )
                
                self.account = Account(
                    primary_smtp_address=self.email,
                    config=config,
                    autodiscover=False
                )
                
                logger.info("✓ Successfully connected to Outlook")
                return True
                
            except UnauthorizedError:
                logger.error("Authorization failed. Check email/password")
                return False
            except AutoDiscoverFailed as e:
                logger.warning(f"Autodiscover failed: {e}")
                if attempt < self.max_retries - 1:
                    import time
                    time.sleep(Config.RETRY_DELAY)
            except Exception as e:
                logger.error(f"Outlook connection error: {e}")
                if attempt < self.max_retries - 1:
                    import time
                    time.sleep(Config.RETRY_DELAY)
        
        logger.error("Failed to connect to Outlook after all retries")
        return False
    
    def download_excel_attachments(self) -> List[str]:
        """Download Excel attachments from inbox"""
        
        if not self.account:
            logger.error("Not connected to Outlook")
            return []
        
        downloaded_files = []
        
        try:
            # Create temp directory
            Path(Config.TEMP_DIR).mkdir(exist_ok=True)
            
            # Calculate date filter
            start_date = datetime.now() - timedelta(days=Config.DAYS_BACK)
            
            logger.info(f"Searching for emails from past {Config.DAYS_BACK} days...")
            
            # Get inbox items
            inbox = self.account.inbox
            items = inbox.filter(received__gte=start_date).order_by('-received')[:100]
            
            for item in items:
                for attachment in item.attachments:
                    if isinstance(attachment, FileAttachment):
                        if attachment.name.lower().endswith(('.xlsx', '.xls', '.csv')):
                            try:
                                filename = Path(Config.TEMP_DIR) / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{attachment.name}"
                                
                                with open(filename, 'wb') as f:
                                    f.write(attachment.content)
                                
                                downloaded_files.append(str(filename))
                                logger.info(f"✓ Downloaded: {attachment.name} ({filename})")
                                
                            except Exception as e:
                                logger.error(f"Failed to save attachment {attachment.name}: {e}")
            
            if not downloaded_files:
                logger.info("No Excel files found in recent emails")
            else:
                logger.info(f"Total files downloaded: {len(downloaded_files)}")
            
            return downloaded_files
            
        except Exception as e:
            logger.error(f"Error downloading attachments: {e}")
            return []


# ============================================================================
# EXCEL PROCESSING
# ============================================================================

class ExcelProcessor:
    """Handle Excel file reading and validation"""
    
    @staticmethod
    def read_excel(filepath: str) -> Tuple[List[Dict], bool]:
        """
        Read Excel file and return data as list of dictionaries
        
        Returns:
            Tuple of (data list, success boolean)
        """
        
        try:
            logger.info(f"Reading Excel file: {filepath}")
            
            workbook = load_workbook(filepath, data_only=True)
            sheet = workbook.active
            
            if not sheet:
                logger.error("No active sheet found in workbook")
                return [], False
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            headers = [str(h).strip() if h else f'column_{i}' for i, h in enumerate(headers)]
            
            # Get data rows
            data = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Skip empty rows
                if all(cell is None for cell in row):
                    continue
                
                row_dict = dict(zip(headers, row))
                data.append(row_dict)
            
            logger.info(f"✓ Successfully read {len(data)} rows from Excel file")
            return data, True
            
        except Exception as e:
            logger.error(f"Failed to read Excel file {filepath}: {e}")
            logger.error(traceback.format_exc())
            return [], False


# ============================================================================
# MYSQL DATABASE OPERATIONS
# ============================================================================

class MySQLDatabase:
    """Handle MySQL database operations with connection pooling"""
    
    def __init__(self):
        self.pool = None
        self._init_pool()
    
    def _init_pool(self):
        """Initialize connection pool"""
        try:
            self.pool = pooling.MySQLConnectionPool(
                pool_name='excel_pool',
                pool_size=5,
                pool_reset_session=True,
                host=Config.MYSQL_HOST,
                user=Config.MYSQL_USER,
                password=Config.MYSQL_PASSWORD,
                database=Config.MYSQL_DATABASE,
                port=Config.MYSQL_PORT,
                autocommit=False
            )
            logger.info("✓ MySQL connection pool initialized")
        except MySQLError as e:
            logger.error(f"Failed to create connection pool: {e}")
            raise
    
    def insert_data(self, data: List[Dict], table_name: str) -> int:
        """
        Insert data into MySQL table
        
        Returns:
            Number of successfully inserted rows
        """
        
        if not data:
            logger.warning("No data to insert")
            return 0
        
        connection = None
        cursor = None
        inserted_count = 0
        failed_count = 0
        
        try:
            connection = self.pool.get_connection()
            cursor = connection.cursor()
            
            # Get column info from first row
            columns = list(data[0].keys())
            columns = [col.replace(' ', '_') for col in columns]  # Sanitize column names
            placeholders = ', '.join(['%s'] * len(columns))
            columns_str = ', '.join([f'`{col}`' for col in columns])
            
            # Build INSERT query
            query = f"INSERT INTO `{table_name}` ({columns_str}) VALUES ({placeholders})"
            
            logger.info(f"Inserting data into table: {table_name}")
            
            # Insert rows with error handling
            for row_idx, row in enumerate(data, start=1):
                try:
                    values = tuple(row.get(col.replace('_', ' '), None) for col in columns)
                    cursor.execute(query, values)
                    inserted_count += 1
                    
                except MySQLError as e:
                    failed_count += 1
                    logger.warning(f"Row {row_idx} failed: {e}")
                    # Continue with next row instead of failing entire operation
                    continue
            
            # Commit all changes
            connection.commit()
            
            logger.info(f"✓ Successfully inserted {inserted_count} rows (Failed: {failed_count})")
            return inserted_count
            
        except MySQLError as e:
            logger.error(f"Database error: {e}")
            if connection:
                connection.rollback()
            return 0
        except Exception as e:
            logger.error(f"Unexpected error during insertion: {e}")
            logger.error(traceback.format_exc())
            if connection:
                connection.rollback()
            return 0
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()
    
    def table_exists(self, table_name: str) -> bool:
        """Check if table exists"""
        connection = None
        cursor = None
        
        try:
            connection = self.pool.get_connection()
            cursor = connection.cursor()
            cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
            return cursor.fetchone() is not None
        except Exception as e:
            logger.error(f"Error checking table existence: {e}")
            return False
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()


# ============================================================================
# FILE CLEANUP
# ============================================================================

class FileManager:
    """Handle file operations and cleanup"""
    
    @staticmethod
    def cleanup_temp_files(filepaths: List[str]):
        """Remove temporary files"""
        for filepath in filepaths:
            try:
                path = Path(filepath)
                if path.exists():
                    path.unlink()
                    logger.info(f"✓ Cleaned up: {filepath}")
            except Exception as e:
                logger.warning(f"Failed to delete {filepath}: {e}")


# ============================================================================
# MAIN AUTOMATION CLASS
# ============================================================================

class ExcelToMySQLAutomation:
    """Main automation orchestrator"""
    
    def __init__(self):
        self.outlook = OutlookIntegration(Config.OUTLOOK_EMAIL, Config.OUTLOOK_PASSWORD)
        self.database = MySQLDatabase()
        self.stats = {
            'files_downloaded': 0,
            'files_processed': 0,
            'rows_inserted': 0,
            'errors': []
        }
    
    def run(self) -> bool:
        """Execute the full automation workflow"""
        
        start_time = datetime.now()
        logger.info("=" * 70)
        logger.info("EXCEL TO MYSQL AUTOMATION STARTED")
        logger.info("=" * 70)
        
        success = True
        
        try:
            # Step 1: Connect to Outlook
            if not self.outlook.connect():
                raise Exception("Failed to connect to Outlook")
            
            # Step 2: Download Excel attachments
            files = self.outlook.download_excel_attachments()
            self.stats['files_downloaded'] = len(files)
            
            if not files:
                logger.info("No files to process. Exiting.")
                return True
            
            # Step 3: Process each file
            for filepath in files:
                try:
                    data, success_read = ExcelProcessor.read_excel(filepath)
                    
                    if not success_read or not data:
                        self.stats['errors'].append(f"Failed to read {filepath}")
                        continue
                    
                    # Step 4: Verify table exists
                    if not self.database.table_exists(Config.TABLE_NAME):
                        logger.error(f"Table {Config.TABLE_NAME} does not exist in database")
                        self.stats['errors'].append(f"Table {Config.TABLE_NAME} not found")
                        continue
                    
                    # Step 5: Insert into database
                    inserted = self.database.insert_data(data, Config.TABLE_NAME)
                    self.stats['rows_inserted'] += inserted
                    self.stats['files_processed'] += 1
                    
                except Exception as e:
                    logger.error(f"Error processing {filepath}: {e}")
                    self.stats['errors'].append(str(e))
                    success = False
            
            # Step 6: Cleanup
            FileManager.cleanup_temp_files(files)
            
        except Exception as e:
            logger.error(f"Automation failed: {e}")
            logger.error(traceback.format_exc())
            self.stats['errors'].append(str(e))
            success = False
        
        finally:
            # Generate report
            self._generate_report(start_time, success)
        
        return success
    
    def _generate_report(self, start_time: datetime, success: bool):
        """Generate execution report"""
        
        duration = datetime.now() - start_time
        
        report = f"""
EXCEL TO MYSQL AUTOMATION REPORT
{'='*70}

Status: {'✓ SUCCESS' if success else '✗ FAILED'}
Duration: {duration}

STATISTICS:
- Files Downloaded: {self.stats['files_downloaded']}
- Files Processed: {self.stats['files_processed']}
- Rows Inserted: {self.stats['rows_inserted']}
- Errors: {len(self.stats['errors'])}

{'='*70}
Execution Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
        
        if self.stats['errors']:
            report += "\nERRORS:\n"
            for error in self.stats['errors']:
                report += f"  - {error}\n"
        
        logger.info(report)
        
        # Send notification
        if success:
            EmailNotifier.send_notification(
                f"✓ Excel to MySQL Sync Successful",
                f"Processed {self.stats['files_downloaded']} files, inserted {self.stats['rows_inserted']} rows",
                is_error=False
            )
        else:
            EmailNotifier.send_notification(
                f"✗ Excel to MySQL Sync Failed",
                report,
                is_error=True
            )


# ============================================================================
# ENTRY POINT
# ============================================================================

def main():
    """Main entry point"""
    try:
        Config.validate()
        automation = ExcelToMySQLAutomation()
        success = automation.run()
        sys.exit(0 if success else 1)
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        logger.error(traceback.format_exc())
        sys.exit(1)


if __name__ == "__main__":
    main()
